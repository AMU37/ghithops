import uuid
import openpyxl
from django.core.management.base import BaseCommand
from core.models import Company, Department
from employees.models import Employee, ServiceType
from transport.models import ShiftType

EXCEL_PATH = 'employees.xlsx'
EXCEL_COMPANIES = ['الشركة اليمنية لتكرير السكر', 'شركة راس عيسى', 'الإدارة العامة', 'المشاريع']
DEFAULT_COMPANY_NAME = 'الشركة اليمنية لتكرير السكر'


def generate_code():
    return 'C' + uuid.uuid4().hex[:6].upper()


class Command(BaseCommand):
    help = 'Import employees from employees.xlsx'

    def handle(self, *args, **options):
        self.stdout.write('=== Importing employees from employees.xlsx ===')
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        total_rows = ws.max_row - 1
        self.stdout.write(f'Total rows in Excel: {total_rows}')

        company_map = {c.name: c for c in Company.objects.all()}
        for cname in EXCEL_COMPANIES:
            if cname not in company_map:
                c = Company.objects.create(name=cname, code=generate_code())
                company_map[cname] = c
                self.stdout.write(f'  Created company: {cname}')
            elif not company_map[cname].code:
                company_map[cname].code = generate_code()
                company_map[cname].save(update_fields=['code'])
                self.stdout.write(f'  Fixed code for: {cname}')

        if DEFAULT_COMPANY_NAME in company_map:
            default_company = company_map[DEFAULT_COMPANY_NAME]
        else:
            default_company = list(company_map.values())[0]

        dept_set, shift_set, service_set = set(), set(), set()
        excel_data = []
        for row in range(2, ws.max_row + 1):
            vals = [ws.cell(row, c).value or '' for c in range(1, 11)]
            emp_id = str(vals[0]).strip()
            name = str(vals[1]).strip()
            if not emp_id or not name:
                continue
            phone = str(vals[2]).strip() if vals[2] else ''
            dept = str(vals[3]).strip() if vals[3] else ''
            position = str(vals[4]).strip() if vals[4] else ''
            shift = str(vals[5]).strip() if vals[5] else ''
            service = str(vals[6]).strip() if vals[6] else ''
            city = str(vals[7]).strip() if vals[7] else ''
            status = str(vals[8]).strip() if vals[8] else ''
            company_name = str(vals[9]).strip() if vals[9] else ''
            excel_data.append((emp_id, name, phone, dept, position, shift, service, city, status, company_name))
            if dept: dept_set.add(dept)
            if shift: shift_set.add(shift)
            if service: service_set.add(service)

        self.stdout.write(f'Unique departments: {len(dept_set)}, shift types: {len(shift_set)}, service types: {len(service_set)}')

        existing_depts = set(Department.objects.values_list('name', flat=True))
        for dname in sorted(dept_set):
            if dname not in existing_depts:
                Department.objects.create(name=dname, company=default_company)
                existing_depts.add(dname)

        shift_obj_map = {s.name: s for s in ShiftType.objects.all()}
        for sname in sorted(shift_set):
            if sname not in shift_obj_map:
                s = ShiftType.objects.create(name=sname, company=default_company)
                shift_obj_map[sname] = s

        service_obj_map = {s.name: s for s in ServiceType.objects.all()}
        for svname in sorted(service_set):
            if svname not in service_obj_map:
                s = ServiceType.objects.create(name=svname, company=default_company)
                service_obj_map[svname] = s

        excel_ids = {emp_id for (emp_id, *_) in excel_data}
        old_count = Employee.objects.exclude(employee_id__in=excel_ids).count()
        if old_count:
            deleted, _ = Employee.objects.exclude(employee_id__in=excel_ids).delete()
            self.stdout.write(f'Deleted {old_count} old employees')

        created = 0
        skipped = 0
        dept_cache = {}
        for (emp_id, name, phone, dept, position, shift, service, city, status, company_name) in excel_data:
            if Employee.objects.filter(employee_id=emp_id).exists():
                skipped += 1
                continue

            c = company_map.get(company_name, default_company)

            dept_obj = None
            if dept:
                if dept not in dept_cache:
                    dept_cache[dept] = Department.objects.filter(name=dept).first()
                dept_obj = dept_cache[dept]

            shift_obj = shift_obj_map.get(shift) if shift else None
            service_obj = service_obj_map.get(service) if service else None
            emp_status = 'inactive' if status == 'غير نشط' else 'active'

            Employee.objects.create(
                employee_id=emp_id,
                full_name=name,
                phone=phone or '',
                email='',
                company=c,
                department=dept_obj,
                department_name=dept or '',
                position=position or '',
                shift_type=shift_obj,
                service_type=service_obj,
                city=city or '',
                status=emp_status,
            )
            created += 1

        self.stdout.write(self.style.SUCCESS(
            f'Done! Created: {created}, Skipped: {skipped}, Total: {Employee.objects.count()}'
        ))
